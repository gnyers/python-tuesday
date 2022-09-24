#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''Demonstration of working with spreadsheets

1. Worksheet management of a given spreadsheet:

   1. list the sheets
      CLI args: --list-sheets WORKBOOK
   2. add sheet
      CLI args: --addsheet SHEETNAME WORKBOOK

2. Data management:

   1. dump the data on a sheet to the stdout or a given output file as text
      CLI args: --dump SHEETNAME WORKBOOK
   2. append provided CSV data at the end of a sheet
      CLI args: --append CSVRECORD --sheet SHEETNAME WORKBOOK

3. CLI Interface: as described above
'''

### Import modules
import sys
import os.path
import argparse
import openpyxl

### Constants

DELIM = ';'                            # default delimiter of CSV data

### Function(s) for CLI arg. parsing

def parseargs(cmdline=sys.argv[1:], known_args_only=False):
    p = argparse.ArgumentParser()
    p.add_argument('-w', '--workbook', type=str, required=True,
                   help='File to read from (default: STDIN)')
    sp = p.add_subparsers(help='commands')
    # --- list shees
    sp_list_sheets = sp.add_parser('list', help='List sheets')
    sp_list_sheets.set_defaults(func=list_sheets)
    # --- add sheet
    sp_add_sheet = sp.add_parser('add', help='Add new sheets')
    sp_add_sheet.add_argument('name', help='Name of the new sheet')
    sp_add_sheet.set_defaults(func=add_sheet)
    # --- dump sheet
    sp_dump_sheet = sp.add_parser('dump', help='Dump data of sheet')
    sp_dump_sheet.add_argument('name',
                               help='Name of the sheet to be dumped')
    sp_dump_sheet.set_defaults(func=dump_data)
    # --- add record
    sp_add_rec = sp.add_parser('addrec', help='Add record to sheet')
    sp_add_rec.add_argument('-n', '--name', type=str, default='',
                             help='Name of the sheet to add record to')
    sp_add_rec.add_argument('-d', '--delim', type=str, default=DELIM,
                             help='CSV delimiter')
    sp_add_rec.add_argument('record', type=str,
                             help='Add this record to sheet')
    sp_add_rec.set_defaults(func=add_rec)

    if known_args_only:
        return p.parse_known_args(cmdline)[0] # interested in known args only
    else:
        return p.parse_args(cmdline)          # parse all args!

### Other functions

def wb_open(fname):
    try:
        wb = openpyxl.load_workbook(fname)
    except Exception as e:
        print('Can not open workbook. Aborting', file=sys.stderr)
        print(e, file=sys.stderr)
        sys.exit(10)
    return wb

### Functions implementing requirements

def list_sheets(workbook):
    ''' List the sheets in *workbok*
    '''
    return [ sheet.title for sheet in workbook.worksheets ]

def add_sheet(workbook, name, index=0):
    ''' Create new sheet with *name* in *workbook* at position *index*
    '''
    workbook.create_sheet(title=name, index=index)
    return workbook

def dump_data(sheet):
    ''' Dump all data of *sheet*
    '''
    return [[cell.value for cell in row ] for row in sheet[sheet.dimensions]]

def add_rec(sheet, data):
    ''' Append *data* after the last record in *sheet*
    '''
    # row_nr = len(sheet.row_dimensions)
    last_row_first_cell = list(sheet.iter_rows())[-1][0]
    next_row = last_row_first_cell.row + 1
    next_col = last_row_first_cell.column
    for col, value in enumerate(data, next_col):
        sheet.cell(row=next_row, column=col, value=value)
    return sheet

### main starts here
if __name__ == '__main__':
    args = parseargs()
    wb = wb_open(args.workbook)

    if args.func == list_sheets:
        for s in list_sheets(wb): print(s)

    if args.func == add_sheet:
        wb = add_sheet(workbook=wb, name=args.name)
        wb.save(args.workbook)

    if args.func == dump_data:
        try:
            ws = wb[args.name]
            data = dump_data(ws)
            print(data)
        except KeyError as e:
            print(e.args[0], file=sys.stderr)
            sys.exit(20)

    if args.func == add_rec:
        try:
            rec = args.record.split(args.delim)
            ws = wb[args.name] if args.name else wb.active
            add_rec(ws, rec)
            wb.save(args.workbook)
            print('Added record to sheet: {}'.format(ws.title))
        except KeyError as e:
            print(e.args[0], file=sys.stderr)
            sys.exit(30)

